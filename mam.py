import xlrenderer
import pandas as pd
import numpy as np
import openpyxl
from  fix_border import style_range, patch_worksheet
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

FILE1 = 'C:/Users/Anton.Gorynia/Downloads/tmp/мама/Батоцыренов Б.В. октябрь.xlsx'
NAMES1 = ['Num', 'Date', 'City', 'Des', 'Obj', 'Time', 'Q', 'Code', 'Price']
FILE2 = 'C:/Users/Anton.Gorynia/Downloads/tmp/мама/in 2.xlsx'
NAMES2 = ['Name', 'DateFrom', 'From', 'To', 'Num', 'TimeFrom', 'DateTo', 'Des']
TEMP = 'C:/Users/Anton.Gorynia/Downloads/tmp/мама/template.xlsx'

def readnaryad(path, names, skiprows =0):
#    rb = xlrd.open_workbook('C:/Users/Anton.Gorynia/Downloads/tmp/мама/Батоцыренов Б.В. октябрь.xlsx')
#    sheet = rb.sheet_by_index(0)
    frame = pd.read_excel(path,
                          sheet_name='наряд1', header=None, index= True, skiprows=skiprows, names=names)
    frame = frame[ (frame.Price > 0) & (frame.Des)]
    return frame


def readframe2(path, names, skiprows =0):
    frame = pd.read_excel(path,
                          sheet_name='Sheet1', header=None, index= True, skiprows=skiprows, names=names)

    return frame

def buisnes_trip_report(template,frame1, frame2, row_numb):
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    ws['AU17'].value = frame1.loc[row_numb]['City']
    ws['CG17'].value = frame2.loc[row_numb]['DateFrom']
    ws['CR17'].value = frame2.loc[row_numb]['DateTo']
    delta = frame2.loc[0]['DateTo'] - frame2.loc[row_numb]['DateFrom']
    ws['DC17'].value = delta.days + 1
    ws['CH25'].value = 'Доп. расходы   (суточные)  {} суток'.format(delta.days + 1)
    ws['A21'].value = frame1.loc[row_numb]['Des']
    wb.save('./мама/{} {} {} out.xlsx'.format(frame2.loc[row_numb]['Name'],
                                                                                 frame1.loc[row_numb]['City'],
                                                                                 frame1.loc[row_numb]['Date'].strftime(
                                                                                     "%d-%B")))


frame1 = readnaryad(FILE1, NAMES1, skiprows=17)
frame2 = readframe2(FILE2, NAMES2,  skiprows=1)
row_number = frame1.shape[0]

for row_numb in range(row_number):
    buisnes_trip_report(TEMP, frame1, frame2, row_numb)
