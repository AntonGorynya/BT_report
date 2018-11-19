import sys  # sys нужен для передачи argv в QApplication
from PyQt5 import QtWidgets, QtCore
import design
import pandas as pd
import openpyxl

NAMES1 = ['Num', 'Date', 'City', 'Des', 'Obj', 'Time', 'Q', 'Code', 'Price']
NAMES2 = ['Name', 'DateFrom', 'From', 'To', 'Num', 'TimeFrom', 'DateTo', 'Des']

def readnaryad(path, names, skiprows =0):
#    rb = xlrd.open_workbook('C:/Users/Anton.Gorynia/Downloads/tmp/мама/Батоцыренов Б.В. октябрь.xlsx')
#    sheet = rb.sheet_by_index(0)
    frame = pd.read_excel(path,
                          sheet_name='наряд1', header=None, index= True, skiprows=skiprows, names=names)
    frame = frame[ (frame.Price > 0) & (frame.Des)]
    return frame


def readframe2(path, names, skiprows =0):
    frame = pd.read_excel(path,
                          sheet_name='Sheet1', header=None, index= True, skiprows=skiprows, names=names)

    return frame

def buisnes_trip_report(template,frame1, frame2, row_numb):
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    ws['AU17'].value = frame1.loc[row_numb]['City']
    ws['CG17'].value = frame2.loc[row_numb]['DateFrom']
    ws['CR17'].value = frame2.loc[row_numb]['DateTo']
    delta = frame2.loc[0]['DateTo'] - frame2.loc[row_numb]['DateFrom']
    ws['DC17'].value = delta.days + 1
    ws['CH25'].value = 'Доп. расходы   (суточные)  {} суток'.format(delta.days + 1)
    ws['A21'].value = frame1.loc[row_numb]['Des']
    wb.save('./{} {} {} out.xlsx'.format(frame2.loc[row_numb]['Name'], frame1.loc[row_numb]['City'],
                                                                       frame1.loc[row_numb]['Date'].strftime("%d-%B")))


class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

        self.pushButton.clicked.connect(self.selectFile1)
        self.pushButton_2.clicked.connect(self.selectFile2)
        self.pushButton_3.clicked.connect(self.selectFileT)
        self.pushButton_4.clicked.connect(self.generate_report)



    def selectFile1(self):
        print("choose input file 1")
        file = QtWidgets.QFileDialog.getOpenFileName(self, "Open new file", '.', "(*.xlsx)")
        self.lineEdit.setText("{}".format(file[0]))
        print(file[0])
        return file[0]

    def selectFile2(self):
        print("choose input file 1")
        file = QtWidgets.QFileDialog.getOpenFileName(self, "Open new file", '.', "(*.xlsx)")
        self.lineEdit_2.setText("{}".format(file[0]))
        print(file[0])
        return file[0]

    def selectFileT(self):
        print("choose input file 1")
        file = QtWidgets.QFileDialog.getOpenFileName(self, "Open new file", '.', "(*.xlsx)")
        self.lineEdit_3.setText("{}".format(file[0]))
        print(file[0])
        return file[0]

    def generate_report(self):
        print("ololo")
        naryad = self.lineEdit.text()
        file2 = self.lineEdit_2.text()
        template = self.lineEdit_3.text()

        frame1 = readnaryad(naryad, NAMES1, skiprows=17)
        frame2 = readframe2(file2, NAMES2, skiprows=1)
        row_number = frame1.shape[0]

        for row_numb in range(row_number):
            buisnes_trip_report(template, frame1, frame2, row_numb)


def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение

main()
input("Enter")
#if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
#    main()  # то запускаем функцию main()
#    input("Press Enter")
